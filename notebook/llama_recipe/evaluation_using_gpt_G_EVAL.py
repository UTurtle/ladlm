import os
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openai import OpenAI



TRANING_OUTDIR = "ladlm_v3"



# OpenAI 클라이언트 초기화
client = OpenAI()

api_key = os.getenv('OPENAI_API_KEY')
if not api_key:
    raise ValueError("API key not found. Please set 'OPENAI_API_KEY' in your environment variables.")
else:
    print("API key found.")

# OpenAI API 호출 함수
def generate_prompt_response(explanation, generated_text):
    if not explanation.strip() or not generated_text.strip():
        print("Explanation or Generated Text is empty, skipping.")
        return ""  # Skip empty content

    prompt = (
        f"""You will be given two descriptions of a spectrogram:

        Task Introduction: You will be given two descriptions of a spectrogram. Your task is to rate how well the Generated Description matches the content of the Original Description on one metric.

        Please make sure you read and understand these instructions carefully. Please keep this document open while reviewing, and refer to it as needed.   

        Evaluation Criteria: 
        Similarity (1-5) - Measures how accurately the Generated Description reflects the content of the Original Description.

        5: The Generated Description perfectly matches the Original Description. No content is missing or different.
        4: The Generated Description closely matches the Original Description with very minor differences.
        3: The Generated Description matches adequately but has minor omissions or discrepancies.
        2: The Generated Description matches some aspects but misses significant details.
        1: The Generated Description does not match the Original Description at all.

        Evaluation Steps:
        1. Read both descriptions carefully, noting any key details about frequency ranges, time intervals, and intensity patterns.
        2. Check if all major points in the Original Description (e.g., weaker 4–7 kHz region, intervals of 0.1–0.3 seconds, wind-like sound around 4–8 seconds) are reflected or missing in the Generated Description.
        3. Identify any contradictions or irrelevant additions.
        4. Assign a score from 1 to 5 based on the Evaluation Criteria.
        
        Original Description:
        {explanation}

        Generated Description:
        {generated_text}

        Evaluation Form (scores ONLY):
        - Score:
        - Brief Reasoning:"""

    )

    # prompt = (
    #     f"Explanation About Spectrogram인 {explanation}와 Generated Text인 {generated_text}를 보고 "
    #     f"제대로 생성했는지 한글로 평가하고 마지막으로 점수를 0점에서 10점 사이로 정해, 양식은 다음과 같아:\n\n"
    #     f"Evaluation\n"
    #     f"Does the Generated Text correctly represent the Explanation?: {{평가 내용 작성}}\n"
    #     f"What are the key issues or strengths in the Generated Text?: {{구체적인 문제점이나 장점 작성}}\n"
    #     f"Suggestions for Improvement (if any): {{개선 제안 작성}}\n\n"
    #     f"Score\n"
    #     f"Final Score (0 to 10): {{점수}}"
    # )

    try:
        completion = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": "You are a helpful assistant."},
                {"role": "user", "content": prompt}
            ]
        )
        # Extract the response content
        response_text = completion.choices[0].message.content
        print(response_text)
        return response_text.strip()
    except Exception as e:
        print(f"Failed to generate response. Error: {e}")
        return ""

# 기존 엑셀 파일 로드
input_excel = f"{TRANING_OUTDIR}_compare_table.xlsx"
output_excel = f"{TRANING_OUTDIR}_compare_table_with_prompt_with_G_EVAL.xlsx"

wb = load_workbook(input_excel)
ws = wb.active

# 헤더 동적 추출
headers = [cell.value for cell in ws[1]]

# 열 인덱스 추출
explanation_col = headers.index("Explanation About Spectrogram") + 1
generated_text_col = headers.index("Generated Text") + 1

# 새로운 열 추가 (Prompt Response)
if "Prompt Response" not in headers:
    ws.cell(row=1, column=ws.max_column + 1, value="Prompt Response")
    response_col = ws.max_column
else:
    response_col = headers.index("Prompt Response") + 1

# 데이터 처리 루프
for row in range(2, ws.max_row + 1):
    explanation = ws.cell(row=row, column=explanation_col).value or ""
    generated_text = ws.cell(row=row, column=generated_text_col).value or ""

    # OpenAI API 호출
    response = generate_prompt_response(explanation, generated_text)
    
    # 결과 저장
    ws.cell(row=row, column=response_col, value=response)

# 엑셀 저장
wb.save(output_excel)
print(f"Results saved to {output_excel}")
